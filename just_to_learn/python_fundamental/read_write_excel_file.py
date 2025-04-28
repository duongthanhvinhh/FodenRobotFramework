import openpyxl

book = openpyxl.load_workbook("data/excel_file.xlsx")
sheet = book.active # Refer to the current active sheet

# To switch to another sheet, use two ways below
# sheet1 = book.worksheets[1]
# sheet2 = book["Sheet 2"]

Dict = {}

cell = sheet.cell(row=2, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "FodenEncrypted"
print(sheet.cell(row=2, column=2).value)

book.save("./data/excel_file.xlsx")